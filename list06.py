# coding=utf-8
import openpyxl
import requests
# 获取excel数据
def read_data(filename, sheetname):
    web = openpyxl.load_workbook(filename)
    sheet = web[sheetname]
    max_row = sheet.max_row
    case_list = []
    for i in range(2, max_row+1):
        dict1 = dict(
            id=sheet.cell(row=i, column=1).value,
            url=sheet.cell(row=i, column=5).value,
            data=sheet.cell(row=i, column=6).value,
            expect=sheet.cell(row=i, column=7).value
        )
        case_list.append(dict1)
    return case_list
# 发送请求函数
def api_func(url, res_body):
    requests_header = {"X-Lemonban-Media-Type": "lemonban.v2", "Content-Type": "application/json"}
    res = requests.post(url=url, json=res_body, headers=requests_header)
    res_log = res.json()
    return res_log
# 写入数据保存结果函数
def writer_result(filename, sheetname, row, column, final_result):
    web = openpyxl.load_workbook(filename)
    sheet = web[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    web.save(filename)
# 测试执行函数--把数据拿出来,调用其它函数
def exeture_func(filename, sheetname):
    res = read_data(filename, sheetname)
    for testcase in res:
        case_id = testcase.get('id')
        url = testcase.get('url')
        data = testcase.get('data')
        expect = testcase.get('expect')
        data = eval(data)# 运行被字符串包裹的python表达式
        expect = eval(expect)
        expect_msg = expect.get('msg')
        res_1 = api_func(url=url, res_body=data)
        real_msg = res_1.get('msg')
        print('预期结果为:{}'.format(expect_msg))
        print('实际结果为:{}'.format(real_msg))
        if real_msg == expect_msg:
            print('测试通过')
            final_res = '通过'
        else:
            print('测试不通过')
            final_res = '测试不用过，有bug'
        print('*' * 100)
        writer_result(filename, sheetname, case_id + 1, 8, final_res)

exeture_func('test_case_api.xlsx', 'register')
exeture_func('test_case_api.xlsx', 'login')



